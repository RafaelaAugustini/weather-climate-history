import requests
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

def save_weather_data():
    file_name = "weather_history"
    documentos_path = os.path.join(os.path.expanduser("~"), "Documents") 
    arquivo_path = os.path.join(documentos_path, f"{file_name}.xlsx") 

    if not os.path.exists(arquivo_path): 
        wb = Workbook() 
        planilha = wb.active
        planilha.title = "weather_climate"
        planilha.append(["Data", "Hora", "Umidade (%)", "Temperatura (C°)"])  
        wb.save(arquivo_path) 
        messagebox.showinfo("Info", "Planilha criada com sucesso.")
    else:
        wb = load_workbook(arquivo_path)
        planilha = wb.active
        messagebox.showinfo("Info", f"O arquivo '{file_name}' já existe em '{documentos_path}'")

    latitude = "-23.5489"
    longitude = "-46.6388"
    key = "8a6ad97070f395e1f44bd3e0d4bc55d1"
 
    response = requests.get(f'https://api.openweathermap.org/data/2.5/weather?lat={latitude}&lon={longitude}&appid={key}&units=metric')

    if response.status_code == 200: 
        data = response.json() 
        umidade = data['main']['humidity']
        temperatura = data['main']['temp']
        now = datetime.now()
        data_atual = now.strftime('%Y-%m-%d')
        hora_atual = now.strftime('%H:%M:%S')

        planilha.append([data_atual, hora_atual, umidade, temperatura])
        wb.save(arquivo_path)

        messagebox.showinfo("Sucesso", "Dados do clima adicionados na planilha.")
    else: 
        messagebox.showerror("Erro", f"Erro ao obter dados do clima: {response.status_code}")

root = tk.Tk()
root.title("Histórico do Clima do Tempo")

label = tk.Label(root, text="Clique para obter os dados climáticos", font=("Arial", 14))
label.pack(pady=20)

button = tk.Button(root, text="Coletar Dados", command=save_weather_data, font=("Arial", 12))
button.pack(pady=10)

root.mainloop()
